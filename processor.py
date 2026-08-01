from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# =========================================================
# إعدادات الماستر
# =========================================================

MASTER_SHEET = "CA Data"
MASTER_DATA_START_ROW = 4
MASTER_FACILITY_COLUMN = 1
MASTER_MONTH_COLUMN = 3

CA_ED_SHEET = "CA -ED"
CA_ED_DATA_START_ROW = 3
CA_ED_FACILITY_COLUMN = 1
CA_ED_MONTH_COLUMN = 3

MONTH_OPTIONS = (
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
)

MONTHS = {
    "jan": "january", "january": "january", "يناير": "january",
    "feb": "february", "february": "february", "فبراير": "february",
    "mar": "march", "march": "march", "مارس": "march",
    "apr": "april", "april": "april", "أبريل": "april", "ابريل": "april",
    "may": "may", "مايو": "may",
    "jun": "june", "june": "june", "يونيو": "june",
    "jul": "july", "july": "july", "يوليو": "july",
    "aug": "august", "august": "august", "أغسطس": "august", "اغسطس": "august",
    "sep": "september", "sept": "september", "september": "september", "سبتمبر": "september",
    "oct": "october", "october": "october", "أكتوبر": "october", "اكتوبر": "october",
    "nov": "november", "november": "november", "نوفمبر": "november",
    "dec": "december", "december": "december", "ديسمبر": "december",
}


# =========================================================
# المابينق المباشر
# المصدر: عنوان خلية داخل ملف القسم
# الهدف: عمود فقط داخل الماستر؛ الصف يحدده البرنامج من كود المنشأة والشهر
#
# أضف بقية الخلايا بنفس الشكل:
# ("H6", "G"),
# ("H7", "H"),
# ("H8", "I"),
# =========================================================

AICU_DIRECT_MAPPING = (
    ("h8", "d"),
    ("h10", "e"),
    ("h11", "f"),
    ("h13", "g"),
    ("h15", "h"),
    ("h16", "i"),
    ("h17", "j"),
    ("h18", "k"),
    ("h19", "l"),
    ("h21", "m"),
    ("h22", "n"),
    ("h23", "o"),
    ("h24", "p"),
    ("h26", "q"),
    ("h27", "r"),
)

PICU_DIRECT_MAPPING = (  
    ("g8", "x"),
    ("g10", "y"),
    ("g12", "z"),
    ("g13", "aa"),
    ("g14", "ab"),
    ("g15", "ac"),
    ("g16", "ad"),
    ("g18", "ae"),
    ("g19", "af"),
    ("g20", "ag"),
    ("g21", "ah"),
    
)

NICU_DIRECT_MAPPING = (
    ("h8", "al"),
    ("h9", "am"),
    ("h10", "an"),
    ("h11", "ao"),
    ("h12", "ap"),
    ("h13", "aq"),
    ("h14", "ar"),
    ("h15", "as"),
    ("h16", "at"),
    ("h17", "au"),
    ("h18", "av"),
    ("h19", "aw"),
    ("h20", "ax"),
)

HRP_DIRECT_MAPPING = (
    ("h9", "gu"),
    ("h10", "gv"),
    ("h11", "gw"),
    ("h12", "gx"),
)

MCARE_DIRECT_MAPPING = (
     ("g7", "gz"),
    ("g8", "ha"),
    ("g9", "hb"),
    ("g10", "hc"),
    ("g11", "hd"),
    ("g13", "he"),
    ("g14", "hf"),
    ("g16", "hg"),
    ("g17", "hh"),
    ("g19", "hi"),
    ("g20", "hj"),
    ("g22", "hk"),
    ("g23", "hl"),
    ("g24", "hm"),
    
)

NURSING_DIRECT_MAPPING = (
    ("h8", "gj"),
    ("h9", "gk"),
    ("h10", "gl"),
    ("h11", "gm"),
    ("h12", "gn"),
    ("h13", "go"),
    ("h14", "gp"),
    ("h15", "gq"),
    ("h16", "gr"),
)

# CPR: G17:J22 وعددها 24 خلية، تقابل BG:CD بالترتيب.
CPR_DIRECT_MAPPING = (
    ("G17", "BG"), ("H17", "Bm"), ("I17", "Bs"), ("J17", "By"),
    ("G18", "Bh"), ("H18", "Bn"), ("I18", "Bt"), ("J18", "Bz"),
    ("G19", "Bi"), ("H19", "Bo"), ("I19", "Bu"), ("J19", "ca"),
    ("G20", "Bj"), ("H20", "Bp"), ("I20", "Bv"), ("J20", "cb"),
    ("G21", "Bk"), ("H21", "Bq"), ("I21", "Bw"), ("J21", "cc"),
    ("G22", "bl"), ("H22", "br"), ("I22", "bx"), ("J22", "Cd"),
)

CSMR_DIRECT_MAPPING = (
    ("G13", "CG"), ("G14", "CH"), ("G15", "CI"), ("G16", "CJ"),
    ("H13", "CK"), ("H14", "CL"), ("H15", "CM"), ("H16", "CN"),
    ("I13", "CO"), ("I14", "CP"), ("I15", "CQ"), ("I16", "CR"),
    ("J13", "CS"), ("J14", "CT"), ("J15", "CU"), ("J16", "CV"),
    ("K13", "CW"), ("K14", "CX"), ("K15", "CY"), ("K16", "CZ"),
)

VTE_DIRECT_MAPPING = (
    ("G23", "DI"),
    ("E28", "DJ"), ("E29", "DK"), ("E30", "DL"), ("E31", "DM"),
    ("E32", "DN"), ("E33", "DO"), ("E34", "DP"), ("E35", "DQ"), ("E36", "DR"),
    ("J23", "DX"),
    ("H28", "DY"), ("H29", "DZ"), ("H30", "EA"), ("H31", "EB"),
    ("H32", "EC"), ("H33", "ED"), ("H34", "EE"), ("H35", "EF"), ("H36", "EG"),
    ("M23", "EM"),
    ("K28", "EN"), ("K29", "EO"), ("K30", "EP"), ("K31", "EQ"),
    ("K32", "ER"), ("K33", "ES"), ("K34", "ET"), ("K35", "EU"), ("K36", "EV"),
    ("P23", "FB"),
    ("N28", "FC"), ("N29", "FD"), ("N30", "FE"), ("N31", "FF"),
    ("N32", "FG"), ("N37", "FH"), ("N38", "FI"),
    ("E41", "FN"), ("E42", "FO"), ("E43", "FP"),
    ("E44", "FQ"), ("E45", "FR"), ("E46", "FS"),
)

# CA-UC يستخدم أكثر من شيت، لذلك كل سطر يحتوي: اسم الشيت، خلية المصدر، عمود الهدف.
CA_UC_DIRECT_MAPPING = (
    ("USR", "g2", "D"),
    ("USR", "H2", "E"),
    ("SEPSIS", "H2", "G"), ("SEPSIS", "I2", "H"), ("SEPSIS", "J2", "I"),
    ("SEPSIS", "K2", "J"), ("SEPSIS", "L2", "K"), ("SEPSIS", "M2", "L"), ("SEPSIS", "N2", "M"),
    ("SEPSIS", "H3", "T"), ("SEPSIS", "I3", "U"), ("SEPSIS", "J3", "V"),
    ("SEPSIS", "K3", "W"), ("SEPSIS", "L3", "X"), ("SEPSIS", "M3", "Y"), ("SEPSIS", "N3", "Z"),
    ("AMI", "G2", "AT"), ("AMI", "H2", "AU"), ("AMI", "I2", "AV"),
    ("AMI", "J2", "AW"), ("AMI", "K2", "AX"), ("AMI", "L2", "AY"),
    ("AMI", "M2", "AZ"), ("AMI", "N2", "BA"), ("AMI", "O2", "BB"),
    ("AMI", "P2", "BC"), ("AMI", "Q2", "BD"),
)

DOMAIN_MAPPINGS = {
    "AICU": AICU_DIRECT_MAPPING,
    "PICU": PICU_DIRECT_MAPPING,
    "NICU": NICU_DIRECT_MAPPING,
    "HRP": HRP_DIRECT_MAPPING,
    "MCARE": MCARE_DIRECT_MAPPING,
    "NURSING": NURSING_DIRECT_MAPPING,
    "CPR": CPR_DIRECT_MAPPING,
    "CSMR": CSMR_DIRECT_MAPPING,
}

DOMAIN_KEYWORDS = {
    "AICU": (
        "AICU",
        "ADULT ICU",
        "ADULT INTENSIVE CARE",
    ),

    "CA-UC": (
        "CA-UC",
        "CA--UC",
        "CA_UC",
        "URGENT CARE",
        "CLINICAL AUDIT",
        "URGENT CARE - CLINICAL AUDIT",
    ),

    "CPR": (
        "CPR",
        "CARDIO PULMONARY RESUSCITATION",
    ),

    "CSMR": (
        "CSMR",
        "CAUSE MR",
        "CAUSE MORTALITY RATE",
    ),

    "HRP": (
        "HRP",
        "HIGH RISK PREGNANCY",
    ),

    "MCARE": (
        "MCARE",
        "MATERNITY CARE",
    ),

    "NICU": (
        "NICU",
        "NEONATAL ICU",
    ),

    "NURSING": (
        "NURSING",
    ),

    "PICU": (
        "PICU",
        "PEDIATRIC ICU",
    ),

    "VTE": (
        "VTE",
    ),
}

# =========================================================
# أدوات عامة
# =========================================================

def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def normalize_facility_code(value) -> str:
    text = clean_text(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def normalize_month(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%B").lower()
        except Exception:
            pass

    text = clean_text(value)
    simple = re.sub(r"[^a-z\u0600-\u06FF]+", " ", text)
    simple = re.sub(r"\s+", " ", simple).strip()
    if simple in MONTHS:
        return MONTHS[simple]
    for key, month in MONTHS.items():
        if re.search(rf"\b{re.escape(key)}\b", simple):
            return month
    return simple


def is_value_present(value) -> bool:
    return value is not None and value != ""


BLACK_FILL = PatternFill(fill_type="solid", fgColor="000000")


def normalize_status(value) -> str:
    text = clean_text(value).replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text).strip()

    if text in {"available", "applicable"}:
        return "Available"

    if text in {
        "not available",
        "notavailable",
        "not applicable",
        "notapplicable",
    }:
        return "Not Available"

    if text in {"no cases", "nocases", "no case"}:
        return "No Cases"

    return ""

def normalized_filename(filename: str) -> str:
    return filename.upper().replace("_", "-").replace(" ", "-")


def detect_domain(filename: str) -> str | None:
    name = normalized_filename(filename)
    for domain, keywords in DOMAIN_KEYWORDS.items():
        if any(keyword.upper().replace("_", "-") in name for keyword in keywords):
            return domain
    return None


def detect_file_status(source_path: str, domain: str) -> tuple[str, str, str]:
    workbook = load_workbook(source_path, data_only=True, read_only=True)
    try:
        identity_ws = find_source_sheet(workbook)
        status_cell = "B6" if domain == "VTE" else "B5"
        status = normalize_status(identity_ws[status_cell].value)

        if not status:
            raise ValueError(
                f"الحالة في الخلية [{identity_ws.title}!{status_cell}] غير صحيحة أو فارغة"
            )

        return status, identity_ws.title, status_cell
    finally:
        workbook.close()

def find_source_sheet(workbook):
    for sheet_name in workbook.sheetnames:
        if re.sub(r"\s+", " ", sheet_name.strip().lower()) == "summary sheet":
            return workbook[sheet_name]
    for sheet_name in workbook.sheetnames:
        if "summary" in sheet_name.strip().lower():
            return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def scan_excel_files(folder: str, master_file: str) -> list[str]:
    found: list[str] = []
    master_abs = os.path.abspath(master_file)

    for current_folder, _, filenames in os.walk(folder):
        for filename in filenames:
            lower = filename.lower()
            if filename.startswith("~$") or not lower.endswith((".xlsx", ".xlsm")):
                continue
            full_path = os.path.abspath(os.path.join(current_folder, filename))
            if full_path == master_abs:
                continue
            if "completed" in lower or "import log" in lower:
                continue
            found.append(full_path)

    return sorted(found)


def find_target_row(
    master_ws,
    facility_code,
    month,
    data_start_row: int,
    facility_column: int,
    month_column: int,
) -> int | None:
    wanted_facility = normalize_facility_code(facility_code)
    wanted_month = normalize_month(month)

    if not wanted_facility or not wanted_month:
        return None

    for row in range(data_start_row, master_ws.max_row + 1):
        master_facility = normalize_facility_code(master_ws.cell(row, facility_column).value)
        master_month = normalize_month(master_ws.cell(row, month_column).value)
        if master_facility == wanted_facility and master_month == wanted_month:
            return row

    return None


def apply_direct_mapping(source_ws, master_ws, target_row: int, mapping, status: str) -> dict:
    transferred_cells = 0
    empty_cells = []
    black_cells = []
    zero_cells = []

    for source_cell, target_column in mapping:
        value = source_ws[source_cell].value
        target_cell = f"{target_column}{target_row}"

        if is_value_present(value):
            master_ws[target_cell] = value
            transferred_cells += 1
            continue

        empty_cells.append({
            "sheet": source_ws.title,
            "source_cell": source_cell.upper(),
            "target_cell": target_cell.upper(),
        })

        if status == "Not Applicable":
            master_ws[target_cell].fill = BLACK_FILL
            black_cells.append(target_cell.upper())
        elif status == "No Cases":
            master_ws[target_cell] = 0
            zero_cells.append(target_cell.upper())

    status_warnings = []
    if status == "Applicable" and empty_cells:
        status_warnings.append("Applicable ولكن لا توجد بيانات في الخلايا المحددة")
    elif status == "Not Applicable" and transferred_cells:
        status_warnings.append("Not Applicable ولكن توجد بيانات")

    return {
        "status": status,
        "transferred_cells": transferred_cells,
        "empty_cells": empty_cells,
        "black_cells": black_cells,
        "zero_cells": zero_cells,
        "status_warnings": status_warnings,
    }


def apply_multi_sheet_mapping(source_wb, master_ws, target_row: int, mapping, status: str) -> dict:
    transferred_cells = 0
    empty_cells = []
    black_cells = []
    zero_cells = []

    for sheet_name, source_cell, target_column in mapping:
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"لم يتم العثور على شيت [{sheet_name}] داخل الملف")

        value = source_wb[sheet_name][source_cell].value
        target_cell = f"{target_column}{target_row}"

        if is_value_present(value):
            master_ws[target_cell] = value
            transferred_cells += 1
            continue

        empty_cells.append({
            "sheet": sheet_name,
            "source_cell": source_cell.upper(),
            "target_cell": target_cell.upper(),
        })

        if status == "Not Applicable":
            master_ws[target_cell].fill = BLACK_FILL
            black_cells.append(target_cell.upper())
        elif status == "No Cases":
            master_ws[target_cell] = 0
            zero_cells.append(target_cell.upper())

    status_warnings = []
    if status == "Applicable" and empty_cells:
        status_warnings.append("Applicable ولكن لا توجد بيانات في الخلايا المحددة")
    elif status == "Not Applicable" and transferred_cells:
        status_warnings.append("Not Applicable ولكن توجد بيانات")

    return {
        "status": status,
        "transferred_cells": transferred_cells,
        "empty_cells": empty_cells,
        "black_cells": black_cells,
        "zero_cells": zero_cells,
        "status_warnings": status_warnings,
    }


# =========================================================
# معالجة الملفات
# =========================================================

def read_standard_identity(summary_ws, domain: str):
    if domain == "VTE":
        return (
            summary_ws["C12"].value,
            summary_ws["C13"].value,
            summary_ws["C15"].value,
            summary_ws["C17"].value,
        )
    return (
        summary_ws["C11"].value,
        summary_ws["C12"].value,
        summary_ws["C14"].value,
        summary_ws["C16"].value,
    )


def process_standard_file(
    source_path: str,
    master_ws,
    selected_month: str,
    domain: str,
    status: str,
) -> dict:
    source_wb = load_workbook(
        source_path,
        data_only=True,
        read_only=False,
    )

    try:
        summary_ws = find_source_sheet(source_wb)

        file_month_1, file_month_2, _, facility_code = read_standard_identity(
            summary_ws,
            domain,
        )

        file_month = normalize_month(file_month_1)

        if file_month not in {month.lower() for month in MONTH_OPTIONS}:
            file_month = normalize_month(file_month_2)

        selected_month_normalized = normalize_month(selected_month)

        if not file_month:
            raise ValueError(
                f"تعذر تحديد شهر ملف {domain}"
            )

        if file_month != selected_month_normalized:
            raise ValueError(
                f"شهر الملف [{file_month.title()}] لا يطابق الشهر المختار "
                f"[{selected_month_normalized.title()}]"
            )

        if not normalize_facility_code(facility_code):
            raise ValueError(
                f"كود المنشأة فارغ في ملف {domain}"
            )

        target_row = find_target_row(
            master_ws,
            facility_code,
            selected_month_normalized,
            MASTER_DATA_START_ROW,
            MASTER_FACILITY_COLUMN,
            MASTER_MONTH_COLUMN,
        )

        if target_row is None:
            raise ValueError(
                f"لم يوجد صف مطابق لكود المنشأة [{facility_code}] "
                f"والشهر [{selected_month_normalized}]"
            )

        if domain == "VTE":
            if "VTE" not in source_wb.sheetnames:
                raise ValueError(
                    "لم يتم العثور على شيت VTE داخل الملف"
                )

            mapping_result = apply_direct_mapping(
                source_wb["VTE"],
                master_ws,
                target_row,
                VTE_DIRECT_MAPPING,
                status,
            )

        else:
            mapping_result = apply_direct_mapping(
                summary_ws,
                master_ws,
                target_row,
                DOMAIN_MAPPINGS[domain],
                status,
            )

        return {
            "facility_code": normalize_facility_code(facility_code),
            "target_row": target_row,
            **mapping_result,
        }

    finally:
        source_wb.close()

def process_ca_uc_file(
    source_path: str,
    ca_ed_ws,
    selected_month: str,
    status: str,
) -> dict:

    source_wb = load_workbook(
        source_path,
        data_only=True,
        read_only=False,
    )

    try:
        if "Summary Sheet" in source_wb.sheetnames:
            summary_ws = source_wb["Summary Sheet"]

            file_month_1 = summary_ws["C11"].value
            file_month_2 = summary_ws["C12"].value
            facility_code = summary_ws["C16"].value

        elif "USR" in source_wb.sheetnames:
            raise ValueError(
                "ملف CA-UC لا يحتوي على Summary Sheet للتحقق من الشهر"
            )

        else:
            raise ValueError(
                "لم يتم العثور على Summary Sheet داخل ملف CA-UC"
            )

        # قراءة شهر الملف
        file_month = normalize_month(file_month_1)

        if file_month not in {month.lower() for month in MONTH_OPTIONS}:
            file_month = normalize_month(file_month_2)

        selected_month_normalized = normalize_month(selected_month)

        # التحقق من شهر الملف مقابل الواجهة
        if not file_month:
            raise ValueError(
                "تعذر تحديد الشهر داخل ملف CA-UC"
            )

        if file_month != selected_month_normalized:
            raise ValueError(
                f"شهر ملف CA-UC [{file_month.title()}] لا يطابق "
                f"الشهر المختار في الواجهة "
                f"[{selected_month_normalized.title()}]"
            )

        if not normalize_facility_code(facility_code):
            raise ValueError(
                "كود المنشأة فارغ في ملف CA-UC"
            )

        # هذا يتحقق أيضًا أن الشهر موجود في الماستر
        target_row = find_target_row(
            ca_ed_ws,
            facility_code,
            selected_month_normalized,
            CA_ED_DATA_START_ROW,
            CA_ED_FACILITY_COLUMN,
            CA_ED_MONTH_COLUMN,
        )

        if target_row is None:
            raise ValueError(
                f"لم يوجد صف في شيت [{CA_ED_SHEET}] مطابق "
                f"لكود المنشأة [{facility_code}] "
                f"والشهر [{selected_month_normalized.title()}]"
            )

        mapping_result = apply_multi_sheet_mapping(
            source_wb,
            ca_ed_ws,
            target_row,
            CA_UC_DIRECT_MAPPING,
            status,
        )

        return {
            "facility_code": normalize_facility_code(facility_code),
            "target_row": target_row,
            **mapping_result,
        }

    finally:
        source_wb.close()


# =========================================================
# تحديد الشهر وحفظ الناتج
# =========================================================

def detect_processing_month(folder: str, master_file: str) -> str:
    detected = normalize_month(Path(master_file).stem)
    valid_months = {month.lower() for month in MONTH_OPTIONS}
    if detected in valid_months:
        return detected

    for source_path in scan_excel_files(folder, master_file):
        source_wb = None
        try:
            source_wb = load_workbook(source_path, data_only=True, read_only=True)
            summary_ws = find_source_sheet(source_wb)
            for address in ("C11", "C12"):
                detected = normalize_month(summary_ws[address].value)
                if detected in valid_months:
                    return detected
        except Exception:
            continue
        finally:
            if source_wb is not None:
                source_wb.close()

    raise ValueError(
        "لم أتمكن من تحديد شهر الإدخال تلقائيًا. "
        "ضع اسم الشهر في اسم ملف الماستر أو مرره إلى selected_month."
    )


def unique_output_path(master_file: str) -> str:
    master_path = Path(master_file)
    candidate = master_path.with_name(f"{master_path.stem} - Completed{master_path.suffix}")
    if not candidate.exists():
        return str(candidate)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return str(
        master_path.with_name(
            f"{master_path.stem} - Completed {timestamp}{master_path.suffix}"
        )
    )


# =========================================================
# الدالة التي تستدعيها الواجهة
# =========================================================

def process_folder(
    folder: str,
    master_file: str,
    logger=None,
    selected_month: str | None = None,
) -> dict:
    # logger موجود فقط حتى لا تنكسر الواجهة القديمة.
    del logger

    month = (
        normalize_month(selected_month)
        if selected_month
        else detect_processing_month(folder, master_file)
    )
    if month not in {m.lower() for m in MONTH_OPTIONS}:
        raise ValueError(f"شهر الإدخال غير صحيح: {month}")

    excel_files = scan_excel_files(folder, master_file)
    if not excel_files:
        raise ValueError("لم يتم العثور على ملفات Excel داخل المجلد أو المجلدات الفرعية")

    keep_vba = master_file.lower().endswith(".xlsm")
    master_wb = load_workbook(master_file, keep_vba=keep_vba)

    try:
        if MASTER_SHEET not in master_wb.sheetnames:
            raise ValueError(f"لم يتم العثور على شيت [{MASTER_SHEET}] في ملف الماستر")

        master_ws = master_wb[MASTER_SHEET]
        ca_ed_ws = (
            master_wb[CA_ED_SHEET]
            if CA_ED_SHEET in master_wb.sheetnames
            else None
        )

        success_files = []
        failed_files = []
        warnings = []

        for source_path in excel_files:
            filename = os.path.basename(source_path)
            domain = detect_domain(filename)

            if domain is None:
                continue

            try:
                status, status_sheet, status_cell = detect_file_status(source_path, domain)

                if domain == "CA-UC":
                    if ca_ed_ws is None:
                        raise ValueError(
                            f"لم يتم العثور على شيت [{CA_ED_SHEET}] في ملف الماستر"
                        )
                    file_result = process_ca_uc_file(
                        source_path, ca_ed_ws, month, status
                    )
                else:
                    file_result = process_standard_file(
                        source_path, master_ws, month, domain, status
                    )

                empty_cells = file_result["empty_cells"]
                success_files.append({
                    "file": filename,
                    "domain": domain,
                    "status": status,
                    "status_sheet": status_sheet,
                    "status_cell": status_cell,
                    "facility_code": file_result["facility_code"],
                    "target_row": file_result["target_row"],
                    "transferred_cells": file_result["transferred_cells"],
                    "empty_cells_count": len(empty_cells),
                    "black_cells_count": len(file_result["black_cells"]),
                    "zero_cells_count": len(file_result["zero_cells"]),
                })

                for warning_message in file_result["status_warnings"]:
                    warnings.append({
                        "file": filename,
                        "domain": domain,
                        "status": status,
                        "warning": warning_message,
                    })

            except Exception as error:
                failed_files.append({
                    "file": filename,
                    "domain": domain,
                    "error": str(error),
                })
                continue

        output_path = unique_output_path(master_file)
        master_wb.save(output_path)

        return {
            "output_path": output_path,
            "month": month,
            "total_detected_files": len(success_files) + len(failed_files),
            "success_files": success_files,
            "failed_files": failed_files,
            "warnings": warnings,
        }

    finally:
        master_wb.close()