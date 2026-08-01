# -*- coding: utf-8 -*-
"""ADAA Weekly Processor: LB, BB, OR."""

from __future__ import annotations

import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

MONTHS = {
    "1": "Jan", "01": "Jan", "jan": "Jan", "january": "Jan",
    "2": "Feb", "02": "Feb", "feb": "Feb", "february": "Feb",
    "3": "Mar", "03": "Mar", "mar": "Mar", "march": "Mar",
    "4": "Apr", "04": "Apr", "apr": "Apr", "april": "Apr",
    "5": "May", "05": "May", "may": "May",
    "6": "Jun", "06": "Jun", "jun": "Jun", "june": "Jun",
    "7": "Jul", "07": "Jul", "jul": "Jul", "july": "Jul",
    "8": "Aug", "08": "Aug", "aug": "Aug", "august": "Aug",
    "9": "Sep", "09": "Sep", "sep": "Sep", "sept": "Sep", "september": "Sep",
    "10": "Oct", "oct": "Oct", "october": "Oct",
    "11": "Nov", "nov": "Nov", "november": "Nov",
    "12": "Dec", "dec": "Dec", "december": "Dec",
    "يناير": "Jan", "فبراير": "Feb", "مارس": "Mar", "أبريل": "Apr",
    "ابريل": "Apr", "مايو": "May", "يونيو": "Jun", "يوليو": "Jul",
    "أغسطس": "Aug", "اغسطس": "Aug", "سبتمبر": "Sep",
    "أكتوبر": "Oct", "اكتوبر": "Oct", "نوفمبر": "Nov", "ديسمبر": "Dec",
}


def _clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _compact(value) -> str:
    return re.sub(r"[^a-z0-9\u0600-\u06FF]+", "", _clean(value).lower())


def normalize_month(value) -> str:
    if isinstance(value, (datetime, date)):
        return MONTHS[str(value.month)]
    text = _clean(value)
    if not text:
        raise ValueError("قيمة الشهر فارغة")
    compact = _compact(text)
    for key, month in MONTHS.items():
        if _compact(key) == compact:
            return month
    for key, month in MONTHS.items():
        key2 = _compact(key)
        if key2 and key2 in compact:
            return month
    m = re.search(r"(?<!\d)(1[0-2]|0?[1-9])(?!\d)", text)
    if m:
        return MONTHS[str(int(m.group(1)))]
    raise ValueError(f"تعذر معرفة الشهر من القيمة: {text}")


def normalize_week_number(value) -> int:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        week = int(value)
    else:
        m = re.search(r"(\d+)", _clean(value))
        if not m:
            raise ValueError(f"تعذر معرفة رقم الأسبوع من القيمة: {value}")
        week = int(m.group(1))
    if week < 1:
        raise ValueError(f"رقم الأسبوع غير صحيح: {week}")
    return week


def week_from_day(day: int) -> int:
    if not 1 <= day <= 31:
        raise ValueError(f"رقم اليوم غير صحيح: {day}")
    if day <= 7:
        return 1
    if day <= 14:
        return 2
    if day <= 21:
        return 3
    return 4


def extract_date_from_filename(filename: str) -> datetime:
    stem = Path(filename).stem

    # يدعم أسماء الأشهر مثل: 2026-JUN-07
    named_month = re.search(
        r"(?<!\d)(20\d{2})[-_. ]([A-Za-z]+)[-_. ](0?[1-9]|[12]\d|3[01])(?!\d)",
        stem,
        re.IGNORECASE,
    )
    if named_month:
        year = int(named_month.group(1))
        month_text = named_month.group(2)
        day = int(named_month.group(3))
        month_number = datetime.strptime(normalize_month(month_text), "%b").month
        return datetime(year, month_number, day)

    patterns = [
        (r"(?<!\d)(20\d{2})[-_. ](0?[1-9]|1[0-2])[-_. ](0?[1-9]|[12]\d|3[01])(?!\d)", "ymd"),
        (r"(?<!\d)(0?[1-9]|[12]\d|3[01])[-_. ](0?[1-9]|1[0-2])[-_. ](20\d{2})(?!\d)", "dmy"),
        (r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])(?!\d)", "ymd"),
        (r"(?<!\d)(0[1-9]|[12]\d|3[01])(0[1-9]|1[0-2])(20\d{2})(?!\d)", "dmy"),
    ]
    for pattern, order in patterns:
        m = re.search(pattern, stem)
        if not m:
            continue
        a, b, c = map(int, m.groups())
        if order == "ymd":
            year, month, day = a, b, c
        else:
            day, month, year = a, b, c
        try:
            return datetime(year, month, day)
        except ValueError:
            pass
    raise ValueError(f"لم أجد تاريخاً واضحاً في اسم الملف: {Path(filename).name}")


def find_sheet(workbook, keyword: str):
    wanted = _compact(keyword)
    for sheet in workbook.worksheets:
        if _compact(sheet.title) == wanted:
            return sheet
    for sheet in workbook.worksheets:
        if wanted in _compact(sheet.title):
            return sheet
    raise KeyError(f"لم أجد شيت يحتوي على: {keyword}")


def find_master_sheet(workbook, process_name: str):
    wanted = _compact(process_name)
    for sheet in workbook.worksheets:
        if _compact(sheet.title) == wanted:
            return sheet
    for sheet in workbook.worksheets:
        if wanted in _compact(sheet.title):
            return sheet
    raise KeyError(f"لم أجد شيت {process_name} في ملف الماستر")


def detect_process_type(filename: str) -> str:
    name = _compact(Path(filename).stem)
    if "or" in name or "operatingroom" in name or "operationroom" in name:
        return "OR"
    if "bb" in name or "bloodbank" in name or "بنكالدم" in name:
        return "BB"
    if "lb" in name or "laboratory" in name or "lab" in name:
        return "LB"
    raise ValueError(f"تعذر تحديد نوع الملف من الاسم: {Path(filename).name}. يجب أن يحتوي الاسم على LB أو BB أو OR")


def build_week_key(week_number: int, month_value) -> str:
    return f"Week {week_number}-{normalize_month(month_value)}"


def normalize_week_key(value) -> str:
    """يوحد Week 1-June وWeek 1-Jun وWeek 1 June."""
    text = _clean(value)
    week_number = normalize_week_number(text)

    # نحذف كلمة Week ورقم الأسبوع حتى لا يُفهم الرقم على أنه شهر.
    month_text = re.sub(r"(?i)\bweek\b", " ", text)
    month_text = re.sub(r"(?<!\d)\d+(?!\d)", " ", month_text, count=1)
    month_text = re.sub(r"[-_/]+", " ", month_text).strip()

    month = normalize_month(month_text)
    return f"{week_number}-{month}"


def find_target_row(master_sheet, facility_code: str, week_key: str) -> Optional[int]:
    """يبني فهرس الصفوف مرة واحدة ويطابق الشهر الكامل أو المختصر."""
    target_index = getattr(master_sheet, "_adaa_target_index", None)

    if target_index is None:
        target_index = {}
        for row in range(1, master_sheet.max_row + 1):
            code = _compact(master_sheet.cell(row=row, column=1).value)
            week_value = master_sheet.cell(row=row, column=3).value

            if not code or not _clean(week_value):
                continue

            try:
                normalized_week = normalize_week_key(week_value)
            except ValueError:
                continue

            target_index[(code, normalized_week)] = row

        master_sheet._adaa_target_index = target_index

    return target_index.get(
        (_compact(facility_code), normalize_week_key(week_key))
    )


def read_row_values(sheet, start_col: str, end_col: str, row: int) -> list:
    start = column_index_from_string(start_col)
    end = column_index_from_string(end_col)
    return [sheet.cell(row=row, column=col).value for col in range(start, end + 1)]


def write_contiguous(master_sheet, target_row: int, start_col: str, values: list) -> None:
    start = column_index_from_string(start_col)
    for offset, value in enumerate(values):
        master_sheet.cell(row=target_row, column=start + offset).value = value


def write_specific_columns(master_sheet, target_row: int, columns: list[str], values: list) -> None:
    if len(columns) != len(values):
        raise ValueError("عدد أعمدة الوجهة لا يساوي عدد القيم")
    for column, value in zip(columns, values):
        master_sheet.cell(row=target_row, column=column_index_from_string(column)).value = value


def process_lb(source_wb, master_wb, filename: str) -> dict:
    summary = find_sheet(source_wb, "Summary")
    manual = find_sheet(source_wb, "Manual")
    master_sheet = find_master_sheet(master_wb, "Weekly Data")
    facility_code = _clean(summary["C10"].value)
    week_key = build_week_key(normalize_week_number(summary["C25"].value), summary["C8"].value)
    target_row = find_target_row(master_sheet, facility_code, week_key)
    if target_row is None:
        raise LookupError(f"لم أجد صفاً مطابقاً في شيت LB: Facility={facility_code}, {week_key}")
    write_contiguous(master_sheet, target_row, "D", read_row_values(manual, "W", "AF", 3))
    return {"file": Path(filename).name, "type": "LB", "facility_code": facility_code, "week": week_key, "target_sheet": master_sheet.title, "target_row": target_row, "status": "تم النقل"}


def process_bb(source_wb, master_wb, filename: str) -> dict:
    summary = find_sheet(source_wb, "Summary")
    manual = find_sheet(source_wb, "Manual")
    master_sheet = find_master_sheet(master_wb, "Weekly Data")
    facility_code = _clean(summary["C10"].value)
    file_date = extract_date_from_filename(filename)
    week_key = build_week_key(week_from_day(file_date.day), summary["C8"].value)
    target_row = find_target_row(master_sheet, facility_code, week_key)
    if target_row is None:
        raise LookupError(f"لم أجد صفاً مطابقاً في شيت BB: Facility={facility_code}, {week_key}")
    write_contiguous(master_sheet, target_row, "S", read_row_values(manual, "B", "K", 4))
    return {"file": Path(filename).name, "type": "BB", "facility_code": facility_code, "file_date": file_date.strftime("%Y-%m-%d"), "week": week_key, "target_sheet": master_sheet.title, "target_row": target_row, "status": "تم النقل"}


def process_or(source_wb, master_wb, filename: str) -> dict:
    summary = find_sheet(source_wb, "Summary")
    master_sheet = find_master_sheet(master_wb, "OR")
    facility_code = _clean(summary["C10"].value)
    week_key = build_week_key(normalize_week_number(summary["C24"].value), summary["C8"].value)
    target_row = find_target_row(master_sheet, facility_code, week_key)
    if target_row is None:
        raise LookupError(f"لم أجد صفاً مطابقاً في شيت OR: Facility={facility_code}, {week_key}")
    if _compact(facility_code) == _compact("W1-MDC-1"):
        it_sheet = find_sheet(source_wb, "IT")
        values = read_row_values(it_sheet, "AA", "AG", 2)
        write_specific_columns(master_sheet, target_row, ["D", "F", "P", "R", "S", "T", "U"], values)
        source_range = f"{it_sheet.title}!AA2:AG2"
        destination = "D,F,P,R,S,T,U"
    else:
        manual = find_sheet(source_wb, "Manual")
        values = read_row_values(manual, "BH", "BY", 2)
        write_contiguous(master_sheet, target_row, "D", values)
        source_range = f"{manual.title}!BH2:BY2"
        destination = "D:U"
    return {"file": Path(filename).name, "type": "OR", "facility_code": facility_code, "week": week_key, "source": source_range, "destination": destination, "target_sheet": master_sheet.title, "target_row": target_row, "status": "تم النقل"}


def process_adaa_weekly(master_file: str | Path, hospital_files: Iterable[str | Path], output_file: str | Path) -> dict:
    master_file = Path(master_file)
    output_file = Path(output_file)
    hospital_files = [Path(item) for item in hospital_files]
    if not master_file.exists():
        raise FileNotFoundError(f"ملف الماستر غير موجود: {master_file}")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    if master_file.resolve() != output_file.resolve():
        shutil.copy2(master_file, output_file)
    master_wb = load_workbook(output_file)
    processed, errors = [], []
    for source_path in hospital_files:
        try:
            if not source_path.exists():
                raise FileNotFoundError("الملف غير موجود")
            process_type = detect_process_type(source_path.name)
            source_wb = load_workbook(source_path, data_only=True, read_only=True)
            try:
                if process_type == "LB":
                    result = process_lb(source_wb, master_wb, source_path.name)
                elif process_type == "BB":
                    result = process_bb(source_wb, master_wb, source_path.name)
                else:
                    result = process_or(source_wb, master_wb, source_path.name)
                processed.append(result)
            finally:
                source_wb.close()
        except Exception as exc:
            errors.append({"file": source_path.name, "status": "فشل", "error": str(exc)})
    master_wb.save(output_file)
    master_wb.close()
    return {"processed": processed, "errors": errors, "processed_count": len(processed), "error_count": len(errors), "output_file": str(output_file)}


process_adaa = process_adaa_weekly
