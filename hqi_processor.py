from __future__ import annotations

import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


START_ROW = 10
HOSPITAL_COLUMN = 2   # B
MASTER_MONTH_COLUMN = 6   # F
FIRST_DATA_COLUMN = column_index_from_string("G")
LAST_DATA_COLUMN = column_index_from_string("DD")

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


def normalize_text(value: Any) -> str:
    """توحيد النصوص للمطابقة بدون التأثير على القيمة الأصلية."""
    if value is None:
        return ""

    text = str(value).strip()
    text = " ".join(text.split())

    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ى": "ي",
        "ة": "ه",
        "ؤ": "و",
        "ئ": "ي",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return text.casefold()


def normalize_month(value: Any) -> str:
    """توحيد الشهر سواء كان نصًا أو رقمًا."""
    if value is None:
        return ""

    text = str(value).strip()
    normalized = normalize_text(text)

    month_aliases = {
        "1": "يناير",
        "01": "يناير",
        "يناير": "يناير",
        "january": "يناير",
        "jan": "يناير",

        "2": "فبراير",
        "02": "فبراير",
        "فبراير": "فبراير",
        "february": "فبراير",
        "feb": "فبراير",

        "3": "مارس",
        "03": "مارس",
        "مارس": "مارس",
        "march": "مارس",
        "mar": "مارس",

        "4": "ابريل",
        "04": "ابريل",
        "ابريل": "ابريل",
        "أبريل": "ابريل",
        "april": "ابريل",
        "apr": "ابريل",

        "5": "مايو",
        "05": "مايو",
        "مايو": "مايو",
        "may": "مايو",

        "6": "يونيو",
        "06": "يونيو",
        "يونيو": "يونيو",
        "june": "يونيو",
        "jun": "يونيو",

        "7": "يوليو",
        "07": "يوليو",
        "يوليو": "يوليو",
        "july": "يوليو",
        "jul": "يوليو",

        "8": "اغسطس",
        "08": "اغسطس",
        "اغسطس": "اغسطس",
        "أغسطس": "اغسطس",
        "august": "اغسطس",
        "aug": "اغسطس",

        "9": "سبتمبر",
        "09": "سبتمبر",
        "سبتمبر": "سبتمبر",
        "september": "سبتمبر",
        "sep": "سبتمبر",
        "sept": "سبتمبر",

        "10": "اكتوبر",
        "اكتوبر": "اكتوبر",
        "أكتوبر": "اكتوبر",
        "october": "اكتوبر",
        "oct": "اكتوبر",

        "11": "نوفمبر",
        "نوفمبر": "نوفمبر",
        "november": "نوفمبر",
        "nov": "نوفمبر",

        "12": "ديسمبر",
        "ديسمبر": "ديسمبر",
        "december": "ديسمبر",
        "dec": "ديسمبر",
    }

    return month_aliases.get(normalized, normalized)


def is_excel_file(path: Path) -> bool:
    return (
        path.is_file()
        and path.suffix.lower() in ALLOWED_EXTENSIONS
        and not path.name.startswith("~$")
    )


def find_source_sheet(workbook):
    return workbook["Sheet1"]


def build_master_index(master_workbook, selected_month: str):
    """
    إنشاء فهرس لكل صف في الماستر حسب:
    اسم المستشفى في B + الشهر في F.
    """
    selected_month_key = normalize_month(selected_month)
    index: dict[tuple[str, str], list[tuple[Any, int]]] = defaultdict(list)
    hospitals_in_master: set[str] = set()
    months_by_hospital: dict[str, set[str]] = defaultdict(set)

    for worksheet in master_workbook.worksheets:
        for row_number in range(START_ROW, worksheet.max_row + 1):
            hospital_value = worksheet.cell(
                row=row_number,
                column=HOSPITAL_COLUMN,
            ).value

            month_value = worksheet.cell(
                row=row_number,
                column=MASTER_MONTH_COLUMN,
            ).value

            hospital_key = normalize_text(hospital_value)
            month_key = normalize_month(month_value)

            if not hospital_key:
                continue

            hospitals_in_master.add(hospital_key)

            if month_key:
                months_by_hospital[hospital_key].add(month_key)

            if hospital_key and month_key:
                index[(hospital_key, month_key)].append(
                    (worksheet, row_number)
                )

    return (
        index,
        hospitals_in_master,
        months_by_hospital,
        selected_month_key,
    )


def copy_values(
    source_worksheet,
    source_row: int,
    target_worksheet,
    target_row: int,
) -> None:
    """نقل القيم فقط من G إلى DD دون لمس A:F."""
    for column_number in range(
        FIRST_DATA_COLUMN,
        LAST_DATA_COLUMN + 1,
    ):
        target_worksheet.cell(
            row=target_row,
            column=column_number,
        ).value = source_worksheet.cell(
            row=source_row,
            column=column_number,
        ).value


def process_hqi_folder(
    folder: str,
    master_file: str,
    selected_month: str,
) -> dict[str, Any]:
    """
    معالجة ملفات HQI وتحديث ملف الماستر.

    المطابقة:
    - اسم المستشفى: العمود B.
    - الشهر في الماستر: العمود F.
    - بداية القراءة: الصف 10.
    - النقل: G إلى DD فقط.
    """
    source_folder = Path(folder)
    master_path = Path(master_file)

    if not source_folder.exists():
        raise FileNotFoundError("مجلد ملفات HQI غير موجود.")

    if not master_path.exists():
        raise FileNotFoundError("ملف الماستر غير موجود.")

    if not selected_month or not str(selected_month).strip():
        raise ValueError("يجب اختيار الشهر.")

    keep_vba = master_path.suffix.lower() == ".xlsm"

    master_workbook = load_workbook(
        master_path,
        keep_vba=keep_vba,
    )

    (
        master_index,
        hospitals_in_master,
        months_by_hospital,
        selected_month_key,
    ) = build_master_index(
        master_workbook,
        selected_month,
    )

    source_paths = sorted(
        path
        for path in source_folder.rglob("*")
        if is_excel_file(path)
    )

    transferred_files: list[dict[str, Any]] = []
    missing_hospitals: list[dict[str, Any]] = []
    month_mismatches: list[dict[str, Any]] = []
    duplicate_hospitals: list[dict[str, Any]] = []
    skipped_files: list[dict[str, Any]] = []

    transferred_rows = 0
    source_hospital_counter: Counter[str] = Counter()
    first_file_by_hospital: dict[str, str] = {}

    for source_path in source_paths:
        try:
            source_workbook = load_workbook(
                source_path,
                data_only=True,
                read_only=True,
                keep_vba=source_path.suffix.lower() == ".xlsm",
            )

            source_worksheet = find_source_sheet(source_workbook)
            file_transferred_rows = 0
            file_hospitals: set[str] = set()
            file_has_data = False

            for row_number in range(
                START_ROW,
                source_worksheet.max_row + 1,
            ):
                hospital_value = source_worksheet.cell(
                    row=row_number,
                    column=HOSPITAL_COLUMN,
                ).value

                hospital_key = normalize_text(hospital_value)

                if not hospital_key:
                    continue

                file_has_data = True
                hospital_name = str(hospital_value).strip()
                file_hospitals.add(hospital_key)

                source_hospital_counter[hospital_key] += 1

                if hospital_key not in first_file_by_hospital:
                    first_file_by_hospital[hospital_key] = source_path.name
                elif source_hospital_counter[hospital_key] == 2:
                    duplicate_hospitals.append({
                        "hospital": hospital_name,
                        "first_file": first_file_by_hospital[hospital_key],
                        "duplicate_file": source_path.name,
                    })

                matches = master_index.get(
                    (hospital_key, selected_month_key),
                    [],
                )

                if not matches:
                    if hospital_key not in hospitals_in_master:
                        missing_hospitals.append({
                            "file": source_path.name,
                            "sheet": source_worksheet.title,
                            "row": row_number,
                            "hospital": hospital_name,
                        })
                    else:
                        available_months = sorted(
                            months_by_hospital.get(hospital_key, set())
                        )

                        month_mismatches.append({
                            "file": source_path.name,
                            "sheet": source_worksheet.title,
                            "row": row_number,
                            "hospital": hospital_name,
                            "selected_month": selected_month,
                            "available_months": available_months,
                        })

                    continue

                if len(matches) > 1:
                    skipped_files.append({
                        "file": source_path.name,
                        "sheet": source_worksheet.title,
                        "row": row_number,
                        "hospital": hospital_name,
                        "reason": (
                            "يوجد أكثر من صف مطابق لنفس المستشفى "
                            "والشهر في ملف الماستر."
                        ),
                    })
                    continue

                target_worksheet, target_row = matches[0]

                copy_values(
                    source_worksheet,
                    row_number,
                    target_worksheet,
                    target_row,
                )

                file_transferred_rows += 1
                transferred_rows += 1

            source_workbook.close()

            if not file_has_data:
                skipped_files.append({
                    "file": source_path.name,
                    "reason": (
                        "لم يتم العثور على اسم مستشفى في العمود B "
                        "ابتداءً من الصف 10."
                    ),
                })
            elif file_transferred_rows > 0:
                transferred_files.append({
                    "file": source_path.name,
                    "sheet": source_worksheet.title,
                    "hospitals": len(file_hospitals),
                    "transferred_rows": file_transferred_rows,
                })
            else:
                skipped_files.append({
                    "file": source_path.name,
                    "sheet": source_worksheet.title,
                    "reason": "لم يتم نقل أي صف من الملف.",
                })

        except Exception as error:
            skipped_files.append({
                "file": source_path.name,
                "reason": str(error),
            })

    output_path = master_path.with_name(
        f"{master_path.stem} - HQI Updated{master_path.suffix}"
    )

    master_workbook.save(output_path)
    master_workbook.close()

    return {
        "process": "HQI",
        "selected_month": selected_month,
        "output_path": str(output_path),
        "total_files": len(source_paths),
        "transferred_file_count": len(transferred_files),
        "transferred_rows": transferred_rows,
        "transferred_files": transferred_files,
        "missing_hospitals": missing_hospitals,
        "month_mismatches": month_mismatches,
        "duplicate_hospitals": duplicate_hospitals,
        "skipped_files": skipped_files,
    }
